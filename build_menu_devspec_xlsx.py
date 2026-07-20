# -*- coding: utf-8 -*-
"""메뉴 관리 — '개발 기준'만 추린 가이드 → Excel.
페이지(프로토타입)만 봐서는 알 수 없어 반드시 명세로 확정해야 하는 항목만 담는다."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\박지은\Documents\GitHub\qrorder-mockup\메뉴관리_개발기준.xlsx"

BLUE="3182F6"; BLUE_50="E8F3FF"; INK="191F28"; INK_2="4E5968"; SEC_BG="EEF4FF"
CAT_FILL = {
    "노출":      "E3F6F0",
    "상태 로직":  "E8F3FF",
    "자동 처리":  "FFF4E5",
    "검증·입력":  "FDF2F8",
    "연쇄·파괴":  "FEECEC",
    "가격 모델":  "F0F0FF",
    "옵션·재고":  "EEF7E8",
}
THIN=Side(style="thin",color="D6DBE0"); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
H_FONT=Font(name="맑은 고딕",size=10.5,bold=True,color="FFFFFF")
B_FONT=Font(name="맑은 고딕",size=10); B_BOLD=Font(name="맑은 고딕",size=10,bold=True)
SEC_FONT=Font(name="맑은 고딕",size=12,bold=True,color=INK)
TTL_FONT=Font(name="맑은 고딕",size=17,bold=True,color=INK)
SUB_FONT=Font(name="맑은 고딕",size=10.5,color=INK_2)
H_FILL=PatternFill("solid",fgColor=BLUE); SEC_FILL=PatternFill("solid",fgColor=SEC_BG)
TOP_WRAP=Alignment(horizontal="left",vertical="top",wrap_text=True)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)

wb=openpyxl.Workbook()

def header_row(ws,row,headers,widths):
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width=w
    for c,h in enumerate(headers,1):
        cell=ws.cell(row,c,h); cell.font=H_FONT; cell.fill=H_FILL; cell.alignment=CENTER; cell.border=BORDER
    ws.row_dimensions[row].height=26

# ======================================================================
# 표지 / 원칙
# ======================================================================
ws=wb.active; ws.title="표지·원칙"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=24; ws.column_dimensions["C"].width=96
ws["B2"]="메뉴 관리 — 개발 기준 가이드"; ws["B2"].font=TTL_FONT; ws.merge_cells("B2:C2")
ws["B3"]="프로토타입(페이지)만 봐서는 알 수 없어, 개발 시 반드시 기준으로 삼아야 하는 항목만 추렸습니다."
ws["B3"].font=SUB_FONT; ws.merge_cells("B3:C3")

def kv(r,k,v,kfill=BLUE_50):
    ws.cell(r,2,k).font=B_BOLD; ws.cell(r,2).fill=PatternFill("solid",fgColor=kfill)
    ws.cell(r,2).border=BORDER; ws.cell(r,2).alignment=TOP_WRAP
    ws.cell(r,3,v).font=B_FONT; ws.cell(r,3).border=BORDER; ws.cell(r,3).alignment=TOP_WRAP

r=5
kv(r,"작성","2026-05-29 · 박지은 / 소프트먼트"); r+=1
kv(r,"기준 산출물","index.html 프로토타입 · MENUS / OPT_GROUPS / CAT_MENU / SETS / MENU_STOCK / OPT_STOCK"); r+=2

ws.cell(r,2,"제외한 것 (페이지만 봐도 충분)").font=SEC_FONT; r+=1
for t in [
    "필드의 필수 여부 — 화면에 빨간 * 로 표시됨",
    "입력 타입(숫자/텍스트/선택) — 입력란 형태로 드러남",
    "'선택' 라벨이 붙은 항목의 선택 여부",
    "화면에 그대로 보이는 표 컬럼·버튼·탭 구성",
]:
    ws.cell(r,2,"—").alignment=CENTER; ws.cell(r,2).border=BORDER
    ws.cell(r,3,t).font=B_FONT; ws.cell(r,3).border=BORDER; ws.cell(r,3).alignment=TOP_WRAP; r+=1
r+=1
ws.cell(r,2,"포함한 것 (개발 기준)").font=SEC_FONT; r+=1
for t in [
    "손님 노출 규칙 — 어드민에 보이는 것과 손님에게 보이는 것이 다른 지점",
    "상태 판정 로직 — effectiveStatus 우선순위",
    "영업일 기준 자동 처리 — 화면에 드러나지 않는 시점 동작",
    "비자명한 검증·입력 결정 — 차단 vs 경고, 음수 허용, 코드 규칙 등",
    "연쇄·파괴적 동작 — 삭제 시 정리, import 덮어쓰기",
    "가격·옵션·재고 모델 — 산식과 관계",
    "데이터 모델 확정 과제 — DATA_MODEL.md와 어긋나 개발 전 결정이 필요한 것",
]:
    ws.cell(r,2,"✓").alignment=CENTER; ws.cell(r,2).border=BORDER
    ws.cell(r,3,t).font=B_FONT; ws.cell(r,3).border=BORDER; ws.cell(r,3).alignment=TOP_WRAP; r+=1
r+=1
kv(r,"전체 필드 사전","필드 전수(타입·기본값 포함)는 메뉴관리_명세서.xlsx 또는 spec/02-menu.md 참조"); r+=1

# ======================================================================
# 개발 기준 (단일 필터 표)
# ======================================================================
ws=wb.create_sheet("개발 기준"); ws.sheet_view.showGridLines=False
ws["A1"]="개발 기준 — 페이지만으론 알 수 없는 것"; ws["A1"].font=TTL_FONT
ws["A2"]="'분류'로 필터해서 보세요. '개발 기준'이 구현이 따라야 할 확정 사양입니다."
ws["A2"].font=SUB_FONT; ws["A2"].alignment=TOP_WRAP; ws.merge_cells("A2:E2")
headers=["분류","항목","개발 기준 (반드시 이렇게)","페이지만으론 알 수 없는 이유","관련 필드·함수"]
widths=[11,32,48,40,22]
header_row(ws,4,headers,widths)

rows=[
("노출","분류 태그(대/중/소)","손님 화면에 노출하지 않는다. 사장님 어드민의 검색·필터 전용.","어드민에 입력란이 보여 손님에게도 보일 것처럼 오해할 수 있음","catL2 / catL3 / tag3"),
("노출","상품코드","손님 화면에 노출하지 않는다.","어드민 표에 보이는 값이라 노출로 오해 가능","code"),
("노출","숨김 vs 품절","'숨김'은 손님 목록에서 완전히 제외, '품절'은 회색+라벨로 표시(선택 불가). 둘은 다르게 처리.","둘 다 '안 팔림'처럼 보이지만 노출 방식이 다름","status, effectiveStatus"),
("노출","사진 없음","사진이 없으면 손님 화면에서 사진 영역 자체를 렌더하지 않는다(빈 박스 X).","어드민엔 '없음'으로 표시돼 빈 플레이스홀더를 둘 거라 오해 가능","img"),
("노출","포장가 빈 값","포장가가 비어 있으면(null) 매장가로 노출한다. 0원으로 처리하면 안 됨.","빈 값의 fallback 규칙은 화면으로 알 수 없음","takeoutPrice, getTakeoutPrice()"),
("노출","포장가 ↔ 주문 방식","주문 방식에서 '포장 할인'이 꺼져 있으면 입력된 포장가를 손님 화면에 적용하지 않는다.","메뉴 화면만 봐서는 알 수 없는 교차 의존","takeoutPrice ↔ ORDER_CHANNELS.takeoutDiscount"),
("노출","판매 시간(schedule)","설정 시간 밖이면 손님 화면에서 회색 + '판매 시간 아니에요', 선택 불가.","시간대별 노출 변화는 정적 화면으로 알 수 없음","schedule"),
("노출","해피아워(discount)","활성 + 적용 시간대일 때만 손님에게 할인가를 노출.","조건부 가격 노출 규칙","discount"),
("노출","정렬 순서","메뉴·카테고리의 정렬 순서가 곧 손님 화면 노출 순서다(관리용 정렬 아님).","순서가 손님 노출에 직결된다는 점이 비자명","MENUS/CAT_MENU 배열 순서, order"),

("상태 로직","effectiveStatus 우선순위","고정 우선순위로 합성: ① manualSoldOut → ② 재고 0(자동 품절) → ③ 숨김 → ④ 품절 → ⑤ 판매중. 먼저 맞는 것 적용.","여러 상태가 동시에 참일 때 무엇이 이기는지 화면으로 알 수 없음","effectiveStatus()"),
("상태 로직","품절 종류","soldOutKind: temp=일시 품절(준비 후 재판매) / full=완전 품절(오늘 판매 불가). manualSoldOut=true일 때만 의미.","'품절' 한 단어 뒤의 두 의미 구분이 비자명","soldOutKind"),

("자동 처리","영업일 시작 — 품절 해제","새 영업일 시작(03:00) 시 manualSoldOut를 전체 자동 해제한다('오늘 품절' 의미).","다음날 자동으로 풀린다는 시점 동작은 화면에 안 드러남","manualSoldOut"),
("자동 처리","영업일 시작 — 재고 리셋","새 영업일 시작 시 remaining=prepared로 리셋, prepared는 유지, yesterday 값 갱신.","리셋 규칙·유지 항목은 화면으로 알 수 없음","MENU_STOCK"),
("자동 처리","영업일 경계","영업일 경계는 03:00 KST(자정 아님). 매출·재고·품절 리셋 기준일.","자정이 아니라는 점이 비자명","SPEC.md §1.1"),

("검증·입력","상품코드 규칙","접두 영문 1자(M=메뉴/O=옵션/S=세트) + 숫자 5자리 = 6자. 매장 내 유니크이며 메뉴·옵션·세트가 같은 코드 풀을 공유한다.","세 종류가 코드 네임스페이스를 공유한다는 점이 비자명","genCode()"),
("검증·입력","메뉴명 중복","중복을 차단하지 않는다. 경고 후 사장님이 원하면 그대로 저장 허용.","중복=에러로 막을 거라 오해 가능","saveMenuEdit()"),
("검증·입력","포장가 > 매장가","차단하지 않는다. 경고 후 허용(사장님 의도 존중).","보통 막을 거라 오해 가능","saveMenuEdit()"),
("검증·입력","매장가 상한","0 이상, 1억원 미만 정수.","상한값은 화면으로 알 수 없음","price"),
("검증·입력","1회 최대 수량","0 또는 빈값이면 null(무제한)로 저장.","빈값 처리 규칙","maxQty"),
("검증·입력","옵션 추가금 음수","옵션 항목 추가금(price)은 음수를 허용한다(감액).","'추가금'이라 0 이상일 거라 오해 가능","OPT_GROUPS.items[].price"),
("검증·입력","세트 추가금 음수","세트 후보 extraPrice는 음수를 허용한다.","위와 동일","candidates[].extraPrice"),

("연쇄·파괴","메뉴 삭제 정리","메뉴 삭제 시 오늘의 준비량·세트 후보·옵션 연결에서 해당 메뉴 참조를 함께 정리한다.","고아 참조 방지 — 백엔드 캐스케이드 필요","deleteMenu(), bulkDeleteMenus()"),
("연쇄·파괴","엑셀 가져오기","import는 병합이 아니라 기존 메뉴·옵션·카테고리·세트를 전체 덮어쓴다(확인 후).","import가 추가일 거라 오해 가능, 파괴적","importMenuJSON()"),
("연쇄·파괴","카테고리 삭제","소속 메뉴의 cat 처리(미지정 이동 등)는 정책 미확정 — 개발 시 결정 필요.","프로토타입에 동작이 없어 비워둠","(정책 결정 필요)"),

("가격 모델","세트 가격 범위","min = basePrice + Σ(자리별 min(extraPrice)),  max = basePrice + Σ(자리별 max(extraPrice)).","세트 총액 산식은 화면으로 추론 어려움","setPriceRange()"),
("가격 모델","세트 구성","자리(slot)마다 후보 중 1개 선택, isDefault 후보가 기본 선택값.","slot/후보/기본선택 모델이 비자명","SETS.slots[].candidates[]"),

("옵션·재고","옵션↔메뉴 N:M","옵션 그룹은 여러 메뉴에 연결되며, 양방향 참조(OPT_GROUPS.links / MENUS.opt)를 동기화해야 한다.","양방향 참조 동기화 요구는 비자명","links, opt"),
("옵션·재고","옵션 항목 단위 재고","옵션 '항목'도 메뉴처럼 수량 관리·품절 처리가 가능하다.","메뉴 단위만 될 거라 오해 가능","OPT_STOCK, ensureOptExtras()"),
]
row=5
for rec in rows:
    cat=rec[0]
    for c,val in enumerate(rec,1):
        cell=ws.cell(row,c,val); cell.border=BORDER; cell.font=B_FONT
        cell.alignment=CENTER if c==1 else TOP_WRAP
        if c==1:
            cell.fill=PatternFill("solid",fgColor=CAT_FILL.get(cat,"FFFFFF")); cell.font=B_BOLD
        if c==2: cell.font=B_BOLD
    row+=1
ws.auto_filter.ref=f"A4:E{row-1}"
ws.freeze_panes="A5"

# ======================================================================
# 데이터 모델 확정 과제
# ======================================================================
ws=wb.create_sheet("데이터 모델 확정 과제"); ws.sheet_view.showGridLines=False
ws["A1"]="데이터 모델 확정 과제 (개발 전 결정 필요)"; ws["A1"].font=TTL_FONT
ws["A2"]="프로토타입에서 쓰지만 DATA_MODEL.md에 없거나 어긋나, 개발 전 확정이 필요한 항목."
ws["A2"].font=SUB_FONT; ws["A2"].alignment=TOP_WRAP; ws.merge_cells("A2:D2")
header_row(ws,4,["#","항목","프로토타입 현황","권장 처리 / 결정 필요"],[5,28,38,52])
gaps=[
("1","판매 상태 값","status 3값(판매중/품절/숨김)","확정 모델은 available/hidden 2값 + 품절은 manualSoldOut·재고로 파생. 레거시 '품절'은 manualSoldOut=true로 마이그레이션 (SPEC §2.3)"),
("2","분류 태그 catL2/catL3/tag3","메뉴에 owner 전용 3단 태그(손님 미노출)","DATA_MODEL §2.2 Menu에 추가하거나 별도 MenuTag(대/중/소) 엔터티로 정규화"),
("3","품절 종류 soldOutKind","메뉴·옵션 항목 모두 보유(temp/full)","Menu/OptionItem에 sold_out_kind enum 추가"),
("4","옵션 항목 재고/품절","OPT_STOCK + 항목 manualSoldOut","OptionItem 단위 재고·품절 모델 추가 (현 StockItem은 메뉴 단위만)"),
("5","카테고리 노출 옵션","displayMode(grid/list)·paused 보유","MenuCategory에 display_mode·paused 추가"),
("6","세트 구조","slot+candidate+extraPrice+isDefault 모델","DATA_MODEL §2.8/2.9의 단순 SetMenuItem(menu_id,quantity)을 자리/후보/추가금 모델로 갱신"),
("7","원산지","자유 텍스트(STORE_ORIGIN.text)","Origin(ingredient, origin) 구조화 모델로 전환 + 마이그레이션"),
("8","상품코드 형식","접두(M/O/S)+숫자 5자리=6자","공통 검증 '영문 대문자+숫자 6자'와 일치. 접두 규칙 명문화 권장"),
]
row=5
for rec in gaps:
    for c,val in enumerate(rec,1):
        cell=ws.cell(row,c,val); cell.border=BORDER; cell.font=B_FONT
        cell.alignment=CENTER if c==1 else TOP_WRAP
        if c==2: cell.font=B_BOLD
    row+=1
ws.auto_filter.ref=f"A4:D{row-1}"
ws.freeze_panes="A5"

wb.save(OUT)
print("SAVED:",OUT); print("SHEETS:",wb.sheetnames); print("기준 건수:",len(rows),"· 확정 과제:",len(gaps))
