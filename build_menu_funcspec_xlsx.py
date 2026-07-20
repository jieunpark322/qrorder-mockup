# -*- coding: utf-8 -*-
"""메뉴 관리 — 화면 기능·동작 명세 → Excel (개발 인계용).

명세서(필드 중심)·개발기준(숨은 규칙)이 다루지 않는 '화면에 보이는 탭·버튼·표 컬럼과
각 동작의 흐름·검증 메시지'를 화면/동작 단위로 정리한다. index.html 프로토타입 기준.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\박지은\OneDrive - (주)소프트먼트\소프트먼트 - 사업부\박지은\Claude\QR오더_사장님관리자_뎁스구조.xlsx"
OUT = r"C:\Users\박지은\Documents\GitHub\qrorder-mockup\메뉴판관리_화면기능명세.xlsx"

# ---- 색/스타일 토큰 (index.html 디자인 토큰과 동일) ----
BLUE = "3182F6"; BLUE_50 = "E8F3FF"; INK = "191F28"; INK_2 = "4E5968"
GREEN_BG = "E3F6F0"; GRAY_BG = "F2F4F6"; BLUE_BG = "E8F3FF"
WARN_BG = "FFF7ED"; PURPLE_BG = "F2EEFF"; SEC_BG = "EEF4FF"

THIN = Side(style="thin", color="D6DBE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H_FONT = Font(name="맑은 고딕", size=10.5, bold=True, color="FFFFFF")
B_FONT = Font(name="맑은 고딕", size=10)
B_BOLD = Font(name="맑은 고딕", size=10, bold=True)
SEC_FONT = Font(name="맑은 고딕", size=12, bold=True, color=INK)
TTL_FONT = Font(name="맑은 고딕", size=17, bold=True, color=INK)
SUB_FONT = Font(name="맑은 고딕", size=10.5, color=INK_2)

H_FILL = PatternFill("solid", fgColor=BLUE)
SEC_FILL = PatternFill("solid", fgColor=SEC_BG)

TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 그룹(탭/화면) 라벨 → 배경색
GROUP_FILL = {
    "공통": GRAY_BG, "메뉴판": BLUE_BG, "옵션 선택지": GREEN_BG,
    "카테고리": WARN_BG, "세트": PURPLE_BG, "원산지": "FEF3F2",
    "메뉴판 관리": BLUE_BG, "오늘의 준비량": GREEN_BG, "번역": WARN_BG,
    "번역(다국어)": WARN_BG,
}


def read_rows(ws, header_row):
    """header_row 다음 행부터 데이터 읽고, 그룹 열(2·3)은 forward-fill."""
    ncol = ws.max_column
    headers = [ws.cell(header_row, c).value for c in range(1, ncol + 1)]
    data, fill2, fill3 = [], "", ""
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
        if all(v in (None, "") for v in vals):
            continue
        if vals[1] not in (None, ""):
            fill2 = vals[1]
        else:
            vals[1] = fill2
        if vals[2] not in (None, ""):
            fill3 = vals[2]
        else:
            vals[2] = fill3
        data.append(["" if v is None else v for v in vals])
    return headers, data


def write_table(ws, top, headers, rows, widths, group_col=2):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ncol = len(headers)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(top, c, h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = H_ALIGN; cell.border = BORDER
    ws.row_dimensions[top].height = 26
    r = top + 1
    for row in rows:
        gname = str(row[group_col - 1]).strip()
        gfill = GROUP_FILL.get(gname)
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            cell.border = BORDER
            cell.font = B_FONT
            if c == 1:
                cell.alignment = CENTER
            else:
                cell.alignment = TOP_WRAP
            if c == group_col:
                cell.font = B_BOLD
                if gfill:
                    cell.fill = PatternFill("solid", fgColor=gfill)
        r += 1
    ws.auto_filter.ref = f"A{top}:{get_column_letter(ncol)}{r - 1}"
    return r


src = openpyxl.load_workbook(SRC, data_only=True)
ov_h, ov_rows = read_rows(src["메뉴관리 기능명세"], 3)
dt_h, dt_rows = read_rows(src["메뉴판관리 상세명세"], 3)

wb = openpyxl.Workbook()

# ============================== 표지 ==============================
ws = wb.active
ws.title = "표지"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 94

ws["B2"] = "QR오더 사장님 어드민 — 메뉴 관리 화면 기능·동작 명세"; ws["B2"].font = TTL_FONT
ws.merge_cells("B2:C2")
ws["B3"] = "화면에 보이는 탭·버튼·표 컬럼과 각 동작의 흐름·검증 메시지 — 개발 인계용"
ws["B3"].font = SUB_FONT; ws.merge_cells("B3:C3")

meta = [
    ("작성일", "2026-05-29"),
    ("작성자", "박지은 / 소프트먼트"),
    ("기준 산출물", "index.html (인터랙티브 프로토타입) — 메뉴 관리 5개 탭"),
    ("관련 문서", "SPEC.md · spec/02-menu.md · 메뉴관리_명세서.xlsx · 메뉴관리_개발기준.xlsx"),
    ("문서 목적", "필드 명세(메뉴관리_명세서)·숨은 규칙(메뉴관리_개발기준)이 다루지 않는 '화면 동작' 계층 — 탭·버튼·표 컬럼·모달 흐름과 각 동작의 검증·확인 문구를 화면/동작 단위로 확정한다."),
]
r = 5
for k, v in meta:
    a = ws.cell(r, 2, k); a.font = B_BOLD; a.fill = PatternFill("solid", fgColor=BLUE_50); a.border = BORDER; a.alignment = TOP_WRAP
    b = ws.cell(r, 3, v); b.font = B_FONT; b.border = BORDER; b.alignment = TOP_WRAP
    r += 1

r += 1
ws.cell(r, 2, "시트 안내").font = SEC_FONT; r += 1
for k, v in [
    ("① 메뉴 관리 기능 개요", "메뉴 관리 전체(메뉴판·옵션·카테고리·세트·원산지·준비량·번역) 기능 한눈에 보기"),
    ("② 메뉴판 화면 기능·동작 상세", "5개 탭의 표 컬럼·버튼·모달·일괄작업과 각 동작의 검증·확인 문구(프로토타입 그대로)"),
]:
    a = ws.cell(r, 2, k); a.font = B_BOLD; a.border = BORDER; a.alignment = TOP_WRAP
    b = ws.cell(r, 3, v); b.font = B_FONT; b.border = BORDER; b.alignment = TOP_WRAP
    r += 1

r += 1
ws.cell(r, 2, "세 문서의 역할 구분").font = SEC_FONT; r += 1
for k, v in [
    ("메뉴관리_명세서", "필드 중심 — 각 값의 타입·필수·기본값·검증·손님 노출"),
    ("메뉴관리_개발기준", "화면만 봐선 모르는 숨은 규칙 — 교차 의존·상태 우선순위·노출 규칙"),
    ("메뉴판관리_화면기능명세 (이 문서)", "화면 동작 중심 — 보이는 탭·버튼·표 컬럼·모달 흐름과 동작별 검증 메시지"),
]:
    a = ws.cell(r, 2, k); a.font = B_BOLD; a.border = BORDER; a.alignment = TOP_WRAP
    b = ws.cell(r, 3, v); b.font = B_FONT; b.border = BORDER; b.alignment = TOP_WRAP
    r += 1

# ===================== ① 메뉴 관리 기능 개요 =====================
ws = wb.create_sheet("① 메뉴 관리 기능 개요")
ws.sheet_view.showGridLines = False
ws["A1"] = "① 메뉴 관리 기능 개요"; ws["A1"].font = TTL_FONT
ws["A2"] = "메뉴 관리 영역 전체 기능을 화면(2depth)·탭(3depth) 단위로 나열. 상세 동작은 ② 시트 참조."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:G2")
write_table(ws, 4, ov_h, ov_rows, [5, 16, 16, 22, 46, 22, 22], group_col=2)
ws.freeze_panes = "A5"

# ================= ② 메뉴판 화면 기능·동작 상세 =================
ws = wb.create_sheet("② 메뉴판 화면 기능·동작 상세")
ws.sheet_view.showGridLines = False
ws["A1"] = "② 메뉴판 화면 기능·동작 상세"; ws["A1"].font = TTL_FONT
ws["A2"] = "메뉴 관리 5개 탭(메뉴판·옵션 선택지·카테고리·세트·원산지)의 표 컬럼·버튼·모달·일괄작업과 동작별 검증·확인 문구. 문구는 프로토타입 그대로."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:G2")
write_table(ws, 4, dt_h, dt_rows, [5, 12, 20, 22, 44, 30, 20], group_col=2)
ws.freeze_panes = "A5"

wb.save(OUT)
print("SAVED:", OUT)
print("SHEETS:", wb.sheetnames)
print("개요 rows:", len(ov_rows), "| 상세 rows:", len(dt_rows))
