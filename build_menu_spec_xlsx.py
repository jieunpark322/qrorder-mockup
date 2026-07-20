# -*- coding: utf-8 -*-
"""메뉴 관리 기능 명세서 + 노출값 기준 → Excel 변환 (개발 인계용)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\박지은\Documents\GitHub\qrorder-mockup\메뉴관리_명세서.xlsx"

# ---- 색/스타일 토큰 (index.html 디자인 토큰과 동일) ----
BLUE     = "3182F6"
BLUE_50  = "E8F3FF"
INK      = "191F28"
INK_2    = "4E5968"
GREEN_BG = "E3F6F0"; GREEN_FG = "0E8C6E"
GRAY_BG  = "F2F4F6"; GRAY_FG  = "8B95A1"
BLUE_BG  = "E8F3FF"; BLUE_FG  = "1D6FE0"
SEC_BG   = "EEF4FF"
WARN_BG  = "FFF7ED"

THIN = Side(style="thin", color="D6DBE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H_FONT   = Font(name="맑은 고딕", size=10.5, bold=True, color="FFFFFF")
B_FONT   = Font(name="맑은 고딕", size=10)
B_BOLD   = Font(name="맑은 고딕", size=10, bold=True)
SEC_FONT = Font(name="맑은 고딕", size=12, bold=True, color=INK)
TTL_FONT = Font(name="맑은 고딕", size=17, bold=True, color=INK)
SUB_FONT = Font(name="맑은 고딕", size=10.5, color=INK_2)

H_FILL   = PatternFill("solid", fgColor=BLUE)
SEC_FILL = PatternFill("solid", fgColor=SEC_BG)

TOP_WRAP   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
H_ALIGN    = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()


def style_expose(cell):
    """손님 노출 셀: 선두 기호(🟢/⚪/🔵)에 따라 색칠."""
    v = str(cell.value or "")
    if v.startswith("🟢"):
        cell.fill = PatternFill("solid", fgColor=GREEN_BG); cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GREEN_FG)
    elif v.startswith("⚪"):
        cell.fill = PatternFill("solid", fgColor=GRAY_BG);  cell.font = Font(name="맑은 고딕", size=10, color=GRAY_FG)
    elif v.startswith("🔵"):
        cell.fill = PatternFill("solid", fgColor=BLUE_BG);  cell.font = Font(name="맑은 고딕", size=10, bold=True, color=BLUE_FG)


def add_table(ws, row, headers, rows, widths, expose_col=None, autofilter=False, first_bold=True):
    """헤더 + 데이터 표를 그린다. expose_col: '손님 노출' 열 인덱스(0-base)."""
    ncol = len(headers)
    # 열 너비
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    # 헤더
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = H_ALIGN; cell.border = BORDER
    ws.row_dimensions[row].height = 26
    hdr_row = row
    row += 1
    # 데이터
    for r in rows:
        for c, val in enumerate(r, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if (ncol > 1 and c in (1,) and len(str(val)) <= 4) else TOP_WRAP
            cell.font = B_FONT
            if first_bold and c == 2:
                cell.font = B_BOLD
            if expose_col is not None and c == expose_col + 1:
                style_expose(cell)
        row += 1
    if autofilter:
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ncol)}{row-1}"
    return row + 1  # 빈 줄


def add_section(ws, row, title, sub=None):
    ws.cell(row=row, column=1, value=title).font = SEC_FONT
    ws.cell(row=row, column=1).fill = SEC_FILL
    # 섹션 배경을 넓게
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = SEC_FILL
    ws.row_dimensions[row].height = 24
    row += 1
    if sub:
        ws.cell(row=row, column=1, value=sub).font = SUB_FONT
        ws.cell(row=row, column=1).alignment = TOP_WRAP
        row += 1
    return row + 0


# ======================================================================
# 표지
# ======================================================================
ws = wb.active
ws.title = "표지"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 92

ws["B2"] = "QR오더 사장님 어드민 — 메뉴 관리 기능 명세서"; ws["B2"].font = TTL_FONT
ws.merge_cells("B2:C2")
ws["B3"] = "기능 상세 명세 · 노출값(손님 노출) 기준 · 데이터 모델 보완 포인트 — 개발 인계용"
ws["B3"].font = SUB_FONT; ws.merge_cells("B3:C3")

meta = [
    ("작성일", "2026-05-29"),
    ("작성자", "박지은 / 소프트먼트"),
    ("기준 산출물", "index.html (인터랙티브 프로토타입) · MENUS / OPT_GROUPS / CAT_MENU / SETS"),
    ("관련 문서", "SPEC.md · DATA_MODEL.md · spec/02-menu.md · spec/03-stock.md · spec/08-common.md"),
    ("문서 목적", "사장님이 어드민에서 입력·관리하는 메뉴 값이 손님 주문 화면에 어떻게 노출되는지를 필드 단위로 확정한다."),
]
r = 5
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = B_BOLD
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=BLUE_50)
    ws.cell(row=r, column=2).border = BORDER; ws.cell(row=r, column=2).alignment = TOP_WRAP
    ws.cell(row=r, column=3, value=v).font = B_FONT
    ws.cell(row=r, column=3).border = BORDER; ws.cell(row=r, column=3).alignment = TOP_WRAP
    r += 1

r += 1
ws.cell(row=r, column=2, value="시트 안내").font = SEC_FONT; r += 1
sheets_info = [
    ("① 메뉴 필드 명세", "메뉴(MENUS) 한 건의 모든 필드 — 타입·필수·기본값·검증·손님 노출"),
    ("② 노출값 기준", "같은 값이 사장님 어드민 ↔ 손님 주문 화면에서 각각 어떻게 보이는지"),
    ("③ 효과적 상태", "effectiveStatus 우선순위 + 손님 화면 노출 규칙"),
    ("④ 옵션 선택지", "옵션 그룹·항목(OPT_GROUPS) 필드 명세"),
    ("⑤ 카테고리", "손님 메뉴판 분류(CAT_MENU) 필드 명세"),
    ("⑥ 세트", "세트 메뉴(SETS) 자리·후보·가격 모델"),
    ("⑦ 원산지·번역", "원산지 표기 / 다국어 요약"),
    ("⑧ 데이터 모델 보완", "프로토타입 ↔ DATA_MODEL.md 차이 · 개발 전 확정 과제 8건"),
]
for k, v in sheets_info:
    ws.cell(row=r, column=2, value=k).font = B_BOLD
    ws.cell(row=r, column=2).border = BORDER; ws.cell(row=r, column=2).alignment = TOP_WRAP
    ws.cell(row=r, column=3, value=v).font = B_FONT
    ws.cell(row=r, column=3).border = BORDER; ws.cell(row=r, column=3).alignment = TOP_WRAP
    r += 1

r += 1
ws.cell(row=r, column=2, value="노출값 범례").font = SEC_FONT; r += 1
legend = [
    ("🟢 노출", "손님 주문 화면(메뉴판·메뉴 상세)에 직접 표시"),
    ("⚪ 미노출", "사장님 어드민에서만 사용. 손님 화면에 표시되지 않음"),
    ("🔵 조건부", "특정 조건에서만 노출되거나, 조건에 따라 노출 형태가 달라짐"),
]
for k, v in legend:
    ck = ws.cell(row=r, column=2, value=k); ck.border = BORDER; ck.alignment = CENTER; style_expose(ck)
    cv = ws.cell(row=r, column=3, value=v); cv.font = B_FONT; cv.border = BORDER; cv.alignment = TOP_WRAP
    r += 1

r += 1
ws.cell(row=r, column=2, value="핵심 구분 — 카테고리 vs 분류 태그").font = SEC_FONT; r += 1
ws.cell(row=r, column=2, value="카테고리(cat)").font = B_BOLD
ws.cell(row=r, column=2).border = BORDER; ws.cell(row=r, column=2).alignment = TOP_WRAP
ws.cell(row=r, column=3, value="🟢 손님 메뉴판의 분류 탭(커피/음료/디저트). 손님에게 보인다.").font = B_FONT
ws.cell(row=r, column=3).border = BORDER; ws.cell(row=r, column=3).alignment = TOP_WRAP; r += 1
ws.cell(row=r, column=2, value="분류 태그(catL2/catL3/tag3)").font = B_BOLD
ws.cell(row=r, column=2).border = BORDER; ws.cell(row=r, column=2).alignment = TOP_WRAP
ws.cell(row=r, column=3, value="⚪ 사장님이 어드민에서 메뉴를 검색·필터할 때 쓰는 대/중/소 자유 키워드. 손님 화면에 절대 노출되지 않는다.").font = B_FONT
ws.cell(row=r, column=3).border = BORDER; ws.cell(row=r, column=3).alignment = TOP_WRAP; r += 1


# ======================================================================
# ① 메뉴 필드 명세
# ======================================================================
ws = wb.create_sheet("① 메뉴 필드 명세")
ws.sheet_view.showGridLines = False
ws["A1"] = "① 메뉴 필드 명세 (MENUS)"; ws["A1"].font = TTL_FONT
ws["A2"] = "메뉴 한 건의 전체 필드. 기본값은 ensureMenuExtras(m)가 최초 접근 시 채운다. '구분' 열로 손님 노출/내부 전용/운영 파생을 필터할 수 있다."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP
ws.merge_cells("A2:I2")

headers = ["구분", "필드", "화면 라벨", "타입", "필수", "기본값", "검증", "손님 노출", "설명 / 비고"]
widths  = [11, 15, 18, 22, 7, 16, 30, 12, 40]
rows = [
    # 손님 노출 값
    ["손님 노출", "name", "메뉴 이름 / 메뉴명", "string(1~30)", "필수", "—", "빈 값 불가. 같은 이름 존재 시 저장 전 확인(차단 X)", "🟢 노출", "메뉴판 카드 제목"],
    ["손님 노출", "price", "매장 판매가", "int(원)", "필수", "0", "0 이상 정수, 1억 미만", "🟢 노출", "매장 식사 시 가격"],
    ["손님 노출", "takeoutPrice", "포장가", "int | null", "선택", "null", "0 이상. 매장가보다 비싸면 저장 전 확인", "🔵 조건부", "포장 채널에서만 적용. null이면 매장가와 동일. 주문 방식에서 포장 할인 OFF면 미적용"],
    ["손님 노출", "cat", "카테고리", "string(CAT_MENU.name)", "필수", "(일괄입력 시 '미지정' 허용)", "CAT_MENU에 존재하는 이름", "🟢 노출", "메뉴판 상단 분류 탭"],
    ["손님 노출", "desc", "설명", "string(~200)", "선택", "''", "—", "🟢 노출", "메뉴명 아래 작은 글씨. 비우면 미표시"],
    ["손님 노출", "img", "메뉴 사진", "string(이모지 | data: | http URL)", "선택", "''", "이미지 JPG·PNG, 5MB 이하", "🔵 조건부", "비우면 손님 화면에 사진 영역 자체가 표시되지 않음"],
    ["손님 노출", "maxQty", "1회 주문 최대 수량", "int | null", "선택", "null(무제한)", "1 이상 정수. 0·빈값이면 null", "🔵 조건부", "담기 수량 상한. 어드민 목록엔 '최대 N개' 배지"],
    ["손님 노출", "opt", "연결된 옵션", "number[] (OPT_GROUPS.id)", "선택", "[]", "존재하는 옵션 그룹 id", "🟢 노출", "메뉴 상세의 옵션 선택지 (④ 시트)"],
    ["손님 노출", "(배열 순서)", "순서", "정렬(드래그)", "—", "등록 순", "—", "🟢 노출", "손님 메뉴판도 여기서 정한 순서 그대로 노출"],
    # 내부 전용
    ["내부 전용", "id", "—", "number", "자동", "자동 부여", "PK", "⚪ 미노출", "내부 식별자"],
    ["내부 전용", "code", "상품코드", "string(6)", "필수(자동생성)", "genCode('M')", "영문 1자(M)+숫자 5자리, 매장 내 유니크(메뉴·옵션·세트 공통 풀)", "⚪ 미노출", "어드민·정산·POS 매칭용. ⚙️로 자동 생성 또는 직접 6자 지정"],
    ["내부 전용", "catL2", "분류 ① (대분류)", "string", "선택", "''", "—", "⚪ 미노출", "어드민 검색·필터 전용. 편집 모달에서 자동완성(datalist) 제안"],
    ["내부 전용", "catL3", "분류 ② (중분류)", "string", "선택", "''", "—", "⚪ 미노출", "동일"],
    ["내부 전용", "tag3", "분류 ③ (소분류)", "string", "선택", "''", "—", "⚪ 미노출", "동일"],
    # 운영·상태 파생
    ["운영·파생", "status", "판매 상태", "enum 판매중 | 품절 | 숨김", "필수", "판매중", "—", "🔵 조건부", "effectiveStatus로 손님 노출 제어 (③ 시트). '품절'은 레거시값 → ⑧ 시트"],
    ["운영·파생", "manualSoldOut", "(오늘 품절 토글)", "bool", "—", "false", "—", "🔵 조건부", "수동 품절. 새 영업일 시작 시 자동 해제"],
    ["운영·파생", "soldOutKind", "(품절 종류)", "enum temp | full", "—", "temp", "—", "🔵 조건부", "temp=일시 품절(준비 후 재판매), full=완전 품절(오늘 판매 불가). manualSoldOut=true일 때만 의미"],
    ["운영·파생", "stock / sold", "(준비량 연동)", "int | null / int", "—", "null / 0", "잔여=stock−sold, 0이면 자동 품절", "🔵 조건부", "오늘의 준비량(MENU_STOCK)과 연동. null이면 수량 무제한 (spec/03-stock.md)"],
    ["운영·파생", "schedule", "판매 시간", "{enabled, days[], start, end}", "선택", "비활성(항상)", "days: 0=일~6=토", "🔵 조건부", "활성 시 시간 외에는 손님 화면에서 회색 + '판매 시간 아니에요'"],
    ["운영·파생", "discount", "해피아워 할인", "{enabled, days[], start, end, rate}", "선택", "비활성", "rate 1~100(%)", "🔵 조건부", "활성 + 시간 충족 시 손님에게 할인가 노출"],
]
add_table(ws, 4, headers, rows, widths, expose_col=7, autofilter=True)
ws.freeze_panes = "A5"


# ======================================================================
# ② 노출값 기준
# ======================================================================
ws = wb.create_sheet("② 노출값 기준")
ws.sheet_view.showGridLines = False
ws["A1"] = "② 노출값 기준 (사장님 어드민 ↔ 손님 주문 화면)"; ws["A1"].font = TTL_FONT
ws["A2"] = "같은 메뉴 데이터가 두 화면에서 어떻게 보이는지의 확정 기준. 손님 앱과 어드민의 렌더링 분기에 사용."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:C2")
headers = ["값", "사장님 어드민", "손님 주문 화면"]
widths  = [22, 30, 60]
rows = [
    ["메뉴명 name", "목록·모달 표시", "🟢 메뉴 카드 제목"],
    ["매장 판매가 price", "표시", "🟢 매장 식사 시 가격"],
    ["포장가 takeoutPrice", "표시(비면 '매장가 동일')", "🔵 포장 채널에서만 적용. null→매장가. 포장 할인 OFF면 매장가로 노출"],
    ["카테고리 cat", "분류 탭·필터", "🟢 메뉴판 상단 분류 탭"],
    ["설명 desc", "표시", "🟢 메뉴명 아래 보조 텍스트(있을 때만)"],
    ["사진 img", "36px 썸네일", "🔵 있을 때만 사진 노출, 없으면 사진 영역 생략"],
    ["1회 최대 수량 maxQty", "'최대 N개' 배지", "🔵 담기 수량 상한 적용"],
    ["연결 옵션 opt", "칩 목록", "🟢 메뉴 상세 옵션 선택지"],
    ["메뉴 순서", "드래그 정렬", "🟢 메뉴판 노출 순서와 동일"],
    ["분류 태그 catL2/catL3/tag3", "드롭다운 필터·검색", "⚪ 노출 안 함"],
    ["상품코드 code", "표시", "⚪ 노출 안 함"],
    ["판매 상태 status", "pill 표시·필터", "🔵 숨김→비표시 / 품절→회색·선택 불가 (③ 시트)"],
    ["오늘 준비량(잔여)", "'남은/준비' + 진행바", "🔵 0이면 '품절'로 노출, 그 외엔 수치 비노출"],
    ["판매 시간 schedule", "표시", "🔵 시간 외 회색 + '판매 시간 아니에요'"],
    ["해피아워 discount", "'🔥 N%↓' 표시", "🔵 적용 시간대에 할인가 노출"],
]
add_table(ws, 4, headers, rows, widths, expose_col=2)
ws.freeze_panes = "A5"


# ======================================================================
# ③ 효과적 상태
# ======================================================================
ws = wb.create_sheet("③ 효과적 상태")
ws.sheet_view.showGridLines = False
ws["A1"] = "③ 메뉴 효과적 상태 (effectiveStatus)"; ws["A1"].font = TTL_FONT
ws["A2"] = "여러 상태 입력을 '지금 손님에게 어떻게 보이는가' 하나로 합성한다. 위→아래 우선순위, 먼저 맞는 것 적용."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:D2")
r = add_table(ws, 4,
    ["우선순위", "조건", "효과적 상태", "비고"],
    [
        ["1", "manualSoldOut = true", "수동 품절", "soldOutKind: temp=일시 / full=완전"],
        ["2", "stock != null AND sold >= stock", "자동 품절", "오늘의 준비량 소진"],
        ["3", "status = '숨김'", "숨김", ""],
        ["4", "status = '품절'", "품절", "레거시값 — ⑧ 시트 참조"],
        ["5", "(그 외)", "판매중", ""],
    ],
    [9, 34, 14, 34], first_bold=False)
ws.cell(row=r, column=1, value="손님 화면 노출 규칙").font = SEC_FONT
for c in range(1, 5): ws.cell(row=r, column=c).fill = SEC_FILL
r += 1
add_table(ws, r,
    ["효과적 상태", "손님 화면"],
    [
        ["판매중 + 판매 시간 충족", "🟢 정상 노출, 선택 가능"],
        ["판매 시간 미충족", "🔵 회색 + '판매 시간 아니에요' (선택 불가)"],
        ["수동/자동 품절", "🔵 회색 + '품절' 라벨 (선택 불가)"],
        ["숨김", "⚪ 목록에서 제외 (아예 표시 안 함)"],
    ],
    [26, 60], expose_col=1, first_bold=True)


# ======================================================================
# ④ 옵션 선택지
# ======================================================================
ws = wb.create_sheet("④ 옵션 선택지")
ws.sheet_view.showGridLines = False
ws["A1"] = "④ 옵션 선택지 (OPT_GROUPS)"; ws["A1"].font = TTL_FONT
ws["A2"] = "메뉴에 붙는 옵션 그룹(예: 샷 추가·사이즈·얼음/온도)과 항목. 한 그룹을 여러 메뉴에 연결(N:M)."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:G2")
headers = ["필드", "화면 라벨", "타입", "필수", "기본값", "손님 노출", "설명"]
widths  = [22, 14, 30, 9, 16, 12, 34]
rows = [
    ["id", "—", "number", "자동", "자동", "⚪ 미노출", ""],
    ["code", "상품코드", "string(6)", "필수(자동)", "genCode('O')", "⚪ 미노출", ""],
    ["name", "옵션 그룹명", "string(~30)", "필수", "—", "🟢 노출", "옵션 그룹 제목"],
    ["mode", "선택 방식", "enum 1개 선택(single) | 중복 선택(multi)", "필수", "1개 선택", "🔵 조건부", "single=라디오, multi=체크박스"],
    ["required", "필수 선택", "bool", "필수", "false", "🔵 조건부", "true면 손님이 반드시 1개 이상 선택"],
    ["items[]", "옵션 항목", "{name, price}[]", "필수(1개↑)", "—", "🟢 노출", "선택지 목록"],
    ["items[].name", "항목명", "string(~30)", "필수", "—", "🟢 노출", ""],
    ["items[].price", "추가 금액", "int(원)", "필수", "0", "🔵 조건부", "음수 허용(감액). 0이면 추가금 표시 안 함"],
    ["links[]", "연결된 메뉴", "number[] (MENUS.id)", "선택", "[]", "🔵 조건부", "연결 메뉴 상세에 노출. 메뉴 쪽은 MENUS.opt로 참조"],
    ["items[].manualSoldOut", "(항목 품절)", "bool", "—", "false", "🔵 조건부", "옵션 항목 단위 수동 품절. 품절 항목은 선택 불가"],
    ["items[].soldOutKind", "(품절 종류)", "enum temp | full", "—", "temp", "🔵 조건부", "오늘의 준비량에서 옵션 항목도 수량 관리·품절 처리 (OPT_STOCK)"],
]
add_table(ws, 4, headers, rows, widths, expose_col=5)
ws.freeze_panes = "A5"


# ======================================================================
# ⑤ 카테고리
# ======================================================================
ws = wb.create_sheet("⑤ 카테고리")
ws.sheet_view.showGridLines = False
ws["A1"] = "⑤ 카테고리 (CAT_MENU)"; ws["A1"].font = TTL_FONT
ws["A2"] = "손님 메뉴판 상단의 분류(탭/구획). 메뉴의 cat 값이 여기를 참조."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:G2")
headers = ["필드", "화면 라벨", "타입", "필수", "기본값", "손님 노출", "설명"]
widths  = [16, 14, 24, 9, 14, 12, 40]
rows = [
    ["id", "—", "number", "자동", "자동", "⚪ 미노출", ""],
    ["name", "카테고리명", "string(~30)", "필수", "—", "🟢 노출", "메뉴판 분류 탭"],
    ["order", "순서", "int / 드래그 정렬", "필수", "등록 순", "🟢 노출", "분류 노출 순서"],
    ["displayMode", "노출 모드", "enum grid | list", "—", "grid", "🔵 조건부", "손님 메뉴판에서 그리드/리스트 형태"],
    ["paused", "운영 중지", "bool", "—", "false", "🔵 조건부", "true면 해당 카테고리·소속 메뉴가 손님 화면에서 빠짐"],
]
add_table(ws, 4, headers, rows, widths, expose_col=5)
ws.cell(row=11, column=1, value="동작: 카테고리 행에서 노출 모드(그리드/리스트)·운영 중지·판매 시간(카테고리 단위) 설정. 카테고리 삭제 시 소속 메뉴 cat 처리(미지정 이동 등)는 정책 확정 필요(⑧).").font = SUB_FONT
ws.cell(row=11, column=1).alignment = TOP_WRAP; ws.merge_cells("A11:G12")
ws.freeze_panes = "A5"


# ======================================================================
# ⑥ 세트
# ======================================================================
ws = wb.create_sheet("⑥ 세트")
ws.sheet_view.showGridLines = False
ws["A1"] = "⑥ 세트 (SETS)"; ws["A1"].font = TTL_FONT
ws["A2"] = "여러 자리(slot)로 구성. 손님은 자리마다 후보 중 하나를 골라 조합을 만든다."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:G2")
headers = ["필드", "화면 라벨", "타입", "필수", "기본값", "손님 노출", "설명"]
widths  = [26, 14, 24, 9, 14, 12, 34]
rows = [
    ["id", "—", "number", "자동", "자동", "⚪ 미노출", ""],
    ["code", "상품코드", "string(6)", "필수(자동)", "genCode('S')", "⚪ 미노출", ""],
    ["name", "세트명", "string(~30)", "필수", "—", "🟢 노출", ""],
    ["basePrice", "기본 가격", "int(원)", "필수", "—", "🔵 조건부", "최저 조합 가격의 기준"],
    ["desc", "설명", "string(~200)", "선택", "''", "🟢 노출", ""],
    ["enabled", "판매 여부", "bool", "필수", "true", "🔵 조건부", "false면 손님 미노출"],
    ["slots[]", "구성 자리", "object[]", "필수(1개↑)", "—", "🟢 노출", "자리별 선택 UI"],
    ["slots[].name", "자리명", "string", "필수", "—", "🟢 노출", "예: '음료 선택'"],
    ["slots[].selectionType", "선택 방식", "enum single", "필수", "single", "🔵 조건부", "자리당 1개 선택"],
    ["slots[].candidates[]", "후보 메뉴", "object[]", "필수(1개↑)", "—", "🟢 노출", "선택지"],
    ["…candidates[].code", "메뉴 코드", "string (MENUS.code)", "필수", "—", "🔵 조건부", "메뉴 참조"],
    ["…candidates[].name", "메뉴명(표시)", "string", "필수", "—", "🟢 노출", ""],
    ["…candidates[].extraPrice", "추가 금액", "int(원)", "선택", "0", "🔵 조건부", "음수 허용. 이 후보 선택 시 가감"],
    ["…candidates[].isDefault", "기본 선택", "bool", "선택", "false", "🔵 조건부", "자리의 기본 선택 후보"],
]
r = add_table(ws, 4, headers, rows, widths, expose_col=5)
ws.cell(row=r, column=1, value="가격 계산 (setPriceRange)").font = SEC_FONT
for c in range(1, 8): ws.cell(row=r, column=c).fill = SEC_FILL
r += 1
ws.cell(row=r, column=1, value="min = basePrice + Σ(자리별 min(extraPrice))\nmax = basePrice + Σ(자리별 max(extraPrice))\n손님: 자리마다 isDefault 후보 기본 선택 → 합계 실시간 표시.  어드민 목록: min~max 범위 노출(단일 조합이면 단일 가격).").font = B_FONT
ws.cell(row=r, column=1).alignment = TOP_WRAP; ws.merge_cells(f"A{r}:G{r+2}")
ws.freeze_panes = "A5"


# ======================================================================
# ⑦ 원산지·번역
# ======================================================================
ws = wb.create_sheet("⑦ 원산지·번역")
ws.sheet_view.showGridLines = False
ws["A1"] = "⑦ 원산지 · 번역(다국어)"; ws["A1"].font = TTL_FONT
add_table(ws, 3,
    ["항목", "프로토타입 현황", "확정 모델 / 노출"],
    [
        ["원산지 (origin)", "STORE_ORIGIN.text — 줄 단위 자유 텍스트. 🟢 손님 화면 풋터에 그대로 노출", "DATA_MODEL.md §2.10 Origin{ingredient, origin, sort_order} 구조화 권장. 마이그레이션: 텍스트를 줄 단위로 파싱해 행 변환 (⑧)"],
        ["번역(다국어)", "카테고리 화면의 언어 미리보기 토글로 손님 화면 다국어 확인. 독립 편집 화면은 후속", "DATA_MODEL.md §2.11 Translation(entity_type, entity_id, locale, field, value). 대상: 메뉴/카테고리/옵션의 name·desc. 번역 없으면 한국어 fallback"],
    ],
    [16, 44, 50], first_bold=True)


# ======================================================================
# ⑧ 데이터 모델 보완
# ======================================================================
ws = wb.create_sheet("⑧ 데이터 모델 보완")
ws.sheet_view.showGridLines = False
ws["A1"] = "⑧ 데이터 모델 보완 포인트 (개발 인계용)"; ws["A1"].font = TTL_FONT
ws["A2"] = "프로토타입에서 쓰지만 DATA_MODEL.md에 아직 없거나, 양쪽이 어긋나 개발 전 확정이 필요한 항목."
ws["A2"].font = SUB_FONT; ws["A2"].alignment = TOP_WRAP; ws.merge_cells("A2:D2")
add_table(ws, 4,
    ["#", "항목", "프로토타입 현황", "권장 처리"],
    [
        ["1", "판매 상태 값", "status 3값(판매중/품절/숨김) 사용", "확정 모델은 available/hidden 2값 + 품절은 manualSoldOut·재고로 파생. 레거시 '품절'은 manualSoldOut=true로 마이그레이션 (SPEC §2.3)"],
        ["2", "분류 태그 catL2/catL3/tag3", "메뉴에 owner 전용 3단 태그 보유, 손님 미노출", "DATA_MODEL §2.2 Menu에 추가하거나 별도 MenuTag(대/중/소) 엔터티로 정규화"],
        ["3", "품절 종류 soldOutKind(temp/full)", "메뉴·옵션 항목 모두 보유", "Menu/OptionItem에 sold_out_kind enum 추가"],
        ["4", "옵션 항목 재고/품절", "OPT_STOCK + 항목 manualSoldOut", "OptionItem 단위 재고·품절 모델 추가 (현 StockItem은 메뉴 단위만)"],
        ["5", "카테고리 노출 옵션 displayMode/paused", "보유", "MenuCategory에 display_mode(grid/list)·paused 추가"],
        ["6", "세트 구조", "slot + candidate + extraPrice + isDefault 모델", "DATA_MODEL §2.8/2.9의 단순 SetMenuItem(menu_id, quantity)을 자리/후보/추가금 모델로 갱신"],
        ["7", "원산지", "자유 텍스트(STORE_ORIGIN.text)", "Origin(ingredient, origin) 구조화 모델로 전환 + 마이그레이션"],
        ["8", "상품코드 형식", "접두 영문 1자(M/O/S)+숫자 5자리 = 6자", "공통 검증 '영문 대문자+숫자 6자'와 일치. 접두 규칙(M=메뉴 등) 명문화 권장"],
    ],
    [5, 26, 36, 50], autofilter=True, first_bold=True)
ws.freeze_panes = "A5"


wb.save(OUT)
print("SAVED:", OUT)
print("SHEETS:", wb.sheetnames)
